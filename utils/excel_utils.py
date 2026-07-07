import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import Config
from db import query


def export_attendance_to_excel(date_from=None, date_to=None, student_id=None):
    """
    Builds an Excel report of attendance records, optionally filtered by
    date range and/or a specific student. Returns the saved file path.
    """
    sql = """
        SELECT s.roll_no, s.name, s.class_section, a.date, a.time_in, a.status
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        WHERE 1=1
    """
    params = []

    if date_from:
        sql += " AND a.date >= %s"
        params.append(date_from)
    if date_to:
        sql += " AND a.date <= %s"
        params.append(date_to)
    if student_id:
        sql += " AND a.student_id = %s"
        params.append(student_id)

    sql += " ORDER BY a.date DESC, s.name ASC"

    rows = query(sql, params, fetch=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Roll No", "Name", "Class/Section", "Date", "Time In", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row["roll_no"],
            row["name"],
            row["class_section"],
            row["date"].strftime("%Y-%m-%d"),
            str(row["time_in"]),
            row["status"],
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    os.makedirs(Config.EXPORT_FOLDER, exist_ok=True)
    filename = f"attendance_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(Config.EXPORT_FOLDER, filename)
    wb.save(filepath)
    return filepath
